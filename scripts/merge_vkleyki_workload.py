"""Merge the new 2026/27 workload workbook without invalidating schedule IDs."""
from __future__ import annotations

import copy, datetime as dt, hashlib, json, re, shutil, sys
from pathlib import Path
from collections import defaultdict
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "timetable_data.json"
HISTORY = ROOT / "data" / "history"

def clean(v): return re.sub(r"\s+", " ", str(v or "")).strip()
def norm(v): return clean(v).casefold().replace("ё", "е")
def valid_teacher(v):
    name = clean(v); low = norm(name)
    return name if name and "вакансия" not in low and low != "вынесена на пп" and len(name.split()) >= 3 else None
def numeric(v):
    try: return float(str(v).replace(",", "."))
    except (TypeError, ValueError): return None
def subgroup(raw, gid):
    if re.search(r"1\s*(?:подгруппа|п/?г)", raw, re.I): return gid * 2
    if re.search(r"2\s*(?:подгруппа|п/?г)", raw, re.I): return gid * 2 + 1
    return -1
def subject_name(raw): return clean(re.sub(r"\s*[12]\s*(?:подгруппа|п/?г)\s*", " ", raw, flags=re.I))
def columns(ws):
    header = subject = teacher = -1
    for r in range(1, min(ws.max_row, 18) + 1):
        for c in range(1, min(ws.max_column, 24) + 1):
            value = norm(ws.cell(r, c).value)
            if "наименование" in value: header, subject = r, c
            if "преподавател" in value: teacher = c
    return header, subject, teacher, 6
def stable_uid(value): return "lesson-" + hashlib.sha256(value.encode()).hexdigest()[:16]

def main():
    source = Path(sys.argv[1])
    data = json.loads(DATA.read_text(encoding="utf-8-sig"))
    schedule = json.loads((ROOT / "output/latest/schedule_all.json").read_text(encoding="utf-8-sig"))
    scheduled_ids = {x["id"] for g in schedule["groups"] for d in g["days"] for s in d["slots"] for x in s.get("lessons", []) if x.get("id", -1) >= 0}
    groups = {norm(x["name"]): x for x in data["groups"]}
    teachers = {norm(x["name"]): x for x in data["teachers"]}
    teacher_names = {x["id"]: x["name"] for x in data["teachers"]}
    group_names = {x["id"]: x["name"] for x in data["groups"]}
    by_exact, by_struct = defaultdict(list), defaultdict(list)
    for old in data["lessons"]:
        parity = old.get("subgroup", -1) % 2 if old.get("subgroup", -1) >= 0 else -1
        base = (norm(group_names[old["group"]]), norm(old["name"]), parity)
        by_struct[base].append(old)
        by_exact[base + (norm(teacher_names.get(old.get("teacher"), "")),)].append(old)

    wb = openpyxl.load_workbook(source, data_only=True)
    rows = []
    for ws in wb.worksheets:
        group = groups.get(norm(ws.title))
        if not group: raise RuntimeError(f"Неизвестная группа: {ws.title}")
        header, scol, tcol, hcol = columns(ws)
        if min(header, scol, tcol) < 1: raise RuntimeError(f"Не найдены колонки: {ws.title}")
        for r in range(header + 1, ws.max_row + 1):
            raw = clean(ws.cell(r, scol).value); hours = numeric(ws.cell(r, hcol).value)
            if not raw or hours is None or hours <= 0: continue
            tname = valid_teacher(ws.cell(r, tcol).value)
            rows.append((group, raw, int(round(hours)), tname))
            if tname and norm(tname) not in teachers:
                new = {"id": max(x["id"] for x in data["teachers"]) + 1, "name": tname,
                       "uid": "teacher-" + hashlib.sha256(norm(tname).encode()).hexdigest()[:16],
                       "work_period":{"from":"","to":""}, "work_days":[{"day":d,"enabled":d<=6,"start_slot":1,"end_slot":7} for d in range(1,8)],
                       "default_room":-1,"campus_priority":[],"room_responsibility":"","availability_note":"",
                       "max_work_days_per_week":0,"max_pairs_per_day":0,"date_slot_overrides":[]}
                data["teachers"].append(new); teachers[norm(tname)] = new

    used = set(); merged = []; updated_teacher = updated_hours = new_count = 0
    next_id = max(x["id"] for x in data["lessons"]) + 1
    for group, raw, hours, tname in rows:
        name = subject_name(raw); sg = subgroup(raw, group["id"]); parity = sg % 2 if sg >= 0 else -1
        base = (norm(group["name"]), norm(name), parity); exact = base + (norm(tname),)
        candidates = [x for x in by_exact.get(exact, []) if x["id"] not in used]
        if not candidates: candidates = [x for x in by_struct.get(base, []) if x["id"] not in used]
        old = candidates[0] if len(candidates) == 1 else None
        tid = teachers[norm(tname)]["id"] if tname else -1
        if old:
            item = copy.deepcopy(old); used.add(item["id"])
            updated_teacher += item.get("teacher") != tid
            updated_hours += item.get("total_hours") != hours
            item.update({"group":group["id"],"subgroup":sg,"teacher":tid,"name":name,"total_hours":hours})
        else:
            item = {"id":next_id,"uid":stable_uid("|".join(map(str, exact))),"group":group["id"],"subgroup":sg,
                    "teacher":tid,"name":name,"total_hours":hours,"total_slots":0,"subject_id":-1,
                    "is_lab":bool(re.search(r"лпз", name, re.I)),"is_block":bool(re.match(r"^(УП|ВУП)\.",name,re.I)),"is_pp":bool(re.match(r"^ПП\.",name,re.I)),
                    "allowed_campuses":[0,1],"week_parity":"all","fixed_room":-1,"preferred_room":-1,
                    "allow_room_substitution":True,"required_room_type":0,"required_capacity":group.get("size",0),
                    "required_equipment":[],"plan_active":False,"generation_active":False}
            next_id += 1; new_count += 1
        merged.append(item)

    # Keep scheduled legacy rows if the workbook no longer contains them; otherwise
    # existing Friday/Saturday references would silently change meaning.
    legacy = [copy.deepcopy(x) for x in data["lessons"] if x["id"] not in used and x["id"] in scheduled_ids]
    merged.extend(legacy)
    HISTORY.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = HISTORY / f"before_vkleyki_merge_{stamp}.json"; shutil.copy2(DATA, backup)
    data["lessons"] = sorted(merged, key=lambda x:x["id"])
    data.setdefault("workload_imports", []).append({"id":max([x.get("id",-1) for x in data.get("workload_imports",[])],default=-1)+1,
        "file_name":source.name,"semester":1,"imported_at":dt.datetime.now(dt.timezone.utc).isoformat(),"groups":len(wb.worksheets),
        "teachers":len(data["teachers"]),"active_lessons":sum(bool(x.get("plan_active",True)) for x in merged),"mode":"stable_id_merge","backup":backup.name})
    DATA.write_text(json.dumps(data, ensure_ascii=False, indent=2)+"\n", encoding="utf-8")
    print(json.dumps({"ok":True,"backup":str(backup),"rows":len(rows),"lessons":len(merged),"legacy_scheduled":len(legacy),
        "new_inactive":new_count,"teachers_updated":updated_teacher,"hours_updated":updated_hours},ensure_ascii=False,indent=2))

if __name__ == "__main__": main()
