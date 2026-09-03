"""Safely replaces groups, teachers and semester-1 workload from the 2026/27 workbook.

Existing settings, rooms and metadata for matching entities are retained. References in
unavailability and substitutions are remapped by stable names. The destination is first
copied to data/history. Run from the repository root:

  python scripts/import_vkleyki_2026.py "C:/path/Вклейки 2026-2027.xlsx"
"""

from __future__ import annotations

import copy
import datetime as dt
import hashlib
import json
import math
import re
import shutil
import sys
from pathlib import Path

import openpyxl

from apply_photo_teacher_availability import (
    DATE_SLOT_OVERRIDES,
    DISABLED_TEACHERS,
    NOTES as CONFIRMED_NOTES,
    RULES as CONFIRMED_RULES,
    UNAVAILABILITY as CONFIRMED_UNAVAILABILITY,
    work_days as confirmed_work_days,
)

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "timetable_data.json"
HISTORY = ROOT / "data" / "history"


def clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm(value) -> str:
    return clean(value).casefold().replace("ё", "е")


def uid(kind: str, value: str) -> str:
    return f"{kind}-{hashlib.sha256(value.encode('utf-8')).hexdigest()[:16]}"


def numeric(value):
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def default_days():
    return [
        {"day": day, "enabled": day <= 6, "start_slot": 1, "end_slot": 7}
        for day in range(1, 8)
    ]


def schedule(enabled_days, start=1, end=7):
    enabled_days = set(enabled_days)
    return [
        {"day": day, "enabled": day in enabled_days, "start_slot": start, "end_slot": end}
        for day in range(1, 8)
    ]


PHOTO_RULES = {
    name: (CONFIRMED_NOTES.get(name, DISABLED_TEACHERS.get(name, "Подтверждено диспетчером")),
           confirmed_work_days(rule))
    for name, rule in CONFIRMED_RULES.items()
}


def valid_teacher(raw) -> str | None:
    name = clean(raw)
    low = norm(name)
    if not name or "вакансия" in low or low == "вынесена на пп":
        return None
    if len(name.split()) < 3:
        return None
    return name


def find_columns(ws):
    header = subject = teacher = index = -1
    for r in range(1, min(ws.max_row, 18) + 1):
        for c in range(1, min(ws.max_column, 24) + 1):
            value = norm(ws.cell(r, c).value)
            if "наименование" in value:
                header, subject = r, c
            if "преподавател" in value:
                teacher = c
            if value == "индекс":
                index = c
    # In this workbook the first-semester workload is column F. The explicit
    # fallback keeps the importer deterministic if merged headers are edited.
    return header, subject, teacher, (index if index > 0 else 2), 6


def subgroup(raw_name: str, group_id: int) -> int:
    if re.search(r"1\s*(?:подгруппа|п/?г)", raw_name, re.I):
        return group_id * 2
    if re.search(r"2\s*(?:подгруппа|п/?г)", raw_name, re.I):
        return group_id * 2 + 1
    return -1


def subject_name(raw_name: str) -> str:
    return clean(re.sub(r"\s*[12]\s*(?:подгруппа|п/?г)\s*", " ", raw_name, flags=re.I))


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Укажите путь к файлу .xlsx")
    source = Path(sys.argv[1])
    if not source.is_file():
        raise SystemExit(f"Файл не найден: {source}")

    current = json.loads(DATA.read_text(encoding="utf-8-sig"))
    old_groups = {norm(x.get("name")): x for x in current.get("groups", [])}
    old_teachers = {norm(x.get("name")): x for x in current.get("teachers", [])}
    old_group_name = {x.get("id"): x.get("name", "") for x in current.get("groups", [])}
    old_teacher_name = {x.get("id"): x.get("name", "") for x in current.get("teachers", [])}
    old_lessons = current.get("lessons", [])

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    HISTORY.mkdir(parents=True, exist_ok=True)
    backup = HISTORY / f"before_vkleyki_import_{stamp}.json"
    shutil.copy2(DATA, backup)

    wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
    groups = []
    for group_id, ws in enumerate(wb.worksheets):
        name = clean(ws.title)
        previous = copy.deepcopy(old_groups.get(norm(name), {}))
        previous.update({"id": group_id, "uid": previous.get("uid") or uid("group", norm(name)), "name": name})
        previous.setdefault("parts", 2)
        previous.setdefault("size", 0)
        previous.setdefault("home_campus", -1)
        previous.setdefault("work_period", {"from": "", "to": ""})
        previous.setdefault("work_days", default_days())
        previous.setdefault("curator_teacher", -1)
        previous.setdefault("class_hour_enabled", True)
        previous.setdefault("class_hour_campus", -1)
        previous.update({
            "class_hour_weekday": 1, "class_hour_slot": 0,
            "class_hour_from": "07:50", "class_hour_to": "09:15",
        })
        groups.append(previous)
    new_group_id = {norm(x["name"]): x["id"] for x in groups}

    # Preserve the workbook's first-appearance order and replace the old roster.
    teacher_names = []
    seen_teachers = set()
    sheet_rows = []
    vacancies = 0
    for ws in wb.worksheets:
        header, subject_col, teacher_col, index_col, hours_col = find_columns(ws)
        if min(header, subject_col, teacher_col) < 1:
            raise RuntimeError(f"{ws.title}: не найдены обязательные колонки")
        rows = []
        for r in range(header + 1, ws.max_row + 1):
            raw_subject = clean(ws.cell(r, subject_col).value)
            hours = numeric(ws.cell(r, hours_col).value)
            if not raw_subject or hours is None or hours <= 0:
                continue
            teacher = valid_teacher(ws.cell(r, teacher_col).value)
            if teacher is None:
                vacancies += 1
            elif norm(teacher) not in seen_teachers:
                seen_teachers.add(norm(teacher))
                teacher_names.append(teacher)
            rows.append((raw_subject, clean(ws.cell(r, index_col).value), int(round(hours)), teacher))
        sheet_rows.append(rows)

    teachers = []
    for teacher_id, name in enumerate(teacher_names):
        previous = copy.deepcopy(old_teachers.get(norm(name), {}))
        previous.update({"id": teacher_id, "uid": previous.get("uid") or uid("teacher", norm(name)), "name": name})
        previous.setdefault("work_period", {"from": "", "to": ""})
        previous.setdefault("work_days", default_days())
        previous.setdefault("default_room", -1)
        previous.setdefault("campus_priority", [])
        previous.setdefault("room_responsibility", "")
        previous.setdefault("availability_note", "")
        previous.setdefault("max_work_days_per_week", 0)
        previous.setdefault("max_pairs_per_day", 0)
        previous.setdefault("date_slot_overrides", [])
        rule = PHOTO_RULES.get(name)
        if rule:
            previous["availability_note"] = rule[0]
            previous["work_days"] = rule[1]
        if name in DISABLED_TEACHERS:
            previous["availability_note"] = DISABLED_TEACHERS[name]
            previous["work_days"] = confirmed_work_days({})
        previous["max_work_days_per_week"] = 1 if name == "Колтышев Евгений Валерьевич" else 0
        previous["max_pairs_per_day"] = 2 if name == "Вагайская Татьяна Александровна" else 0
        if name in DATE_SLOT_OVERRIDES:
            previous["date_slot_overrides"] = [
                {"date": date, "slots": slots}
                for date, slots in sorted(DATE_SLOT_OVERRIDES[name].items())
            ]
        teachers.append(previous)
    teacher_id_by_name = {norm(x["name"]): x["id"] for x in teachers}

    old_lesson_by_key = {}
    for lesson in old_lessons:
        key = (
            norm(old_group_name.get(lesson.get("group"), "")), norm(lesson.get("name")),
            int(lesson.get("subgroup", -1)) % 2 if int(lesson.get("subgroup", -1)) >= 0 else -1,
            norm(old_teacher_name.get(lesson.get("teacher"), "")),
        )
        old_lesson_by_key.setdefault(key, lesson)

    lessons = []
    subject_ids = {}
    next_subject_id = 0
    old_to_new_lesson = {}
    for group, rows in zip(groups, sheet_rows):
        for raw_name, index_value, hours, teacher_name in rows:
            name = subject_name(raw_name)
            teacher = teacher_id_by_name.get(norm(teacher_name), -1)
            sg = subgroup(raw_name, group["id"])
            block = bool(re.match(r"^(?:УП|ВУП)\.", name, re.I))
            subject_key = (norm(group["name"]), norm(index_value or re.sub(r"^ЛПЗ[.\s]+", "", name, flags=re.I)))
            if subject_key not in subject_ids:
                subject_ids[subject_key] = next_subject_id
                next_subject_id += 1
            old_key = (norm(group["name"]), norm(name), sg % 2 if sg >= 0 else -1, norm(teacher_name))
            old = old_lesson_by_key.get(old_key, {})
            lesson_id = len(lessons)
            lesson = {
                "id": lesson_id,
                "uid": old.get("uid") or uid("lesson", "|".join(map(str, old_key))),
                "group": group["id"], "subgroup": sg, "teacher": teacher,
                "name": name, "total_hours": hours,
                "total_slots": max(1, math.ceil(hours / (6 if block else 2))),
                "subject_id": -1 if re.match(r"^(?:КП|УП|ВУП)[.\s]", name, re.I) else subject_ids[subject_key],
                "is_lab": bool(re.match(r"^ЛПЗ[.\s]", name, re.I)),
                "is_block": block, "is_pp": bool(re.match(r"^ПП\.", name, re.I)),
                "allowed_campuses": old.get("allowed_campuses", [0, 1]),
                "week_parity": old.get("week_parity", "all"),
                "fixed_room": old.get("fixed_room", -1),
                "allow_room_substitution": old.get("allow_room_substitution", True),
                "required_room_type": old.get("required_room_type", 1),
                "required_capacity": old.get("required_capacity", group.get("size", 0)),
                "required_equipment": old.get("required_equipment", []),
                "plan_active": True,
            }
            lessons.append(lesson)
            if "id" in old:
                old_to_new_lesson[old["id"]] = lesson_id

    # Curators and relational data follow people/groups by name, never obsolete IDs.
    old_to_new_teacher = {old_id: teacher_id_by_name.get(norm(name), -1) for old_id, name in old_teacher_name.items()}
    old_to_new_group = {old_id: new_group_id.get(norm(name), -1) for old_id, name in old_group_name.items()}
    for group in groups:
        old = old_groups.get(norm(group["name"]), {})
        group["curator_teacher"] = old_to_new_teacher.get(old.get("curator_teacher", -1), -1)

    def remap_items(items, id_field, id_map):
        result = []
        for item in items:
            mapped = id_map.get(item.get(id_field), -1)
            if mapped < 0:
                continue
            clone = copy.deepcopy(item)
            clone[id_field] = mapped
            result.append(clone)
        return result

    current["groups"] = groups
    current["teachers"] = teachers
    current["lessons"] = lessons
    current["unavailable"] = remap_items(current.get("unavailable", []), "group", old_to_new_group)
    current["teacher_unavailable"] = remap_items(current.get("teacher_unavailable", []), "teacher", old_to_new_teacher)

    substitutions = []
    for item in current.get("substitutions", []):
        lesson_id = old_to_new_lesson.get(item.get("lesson_id"), -1)
        absent = old_to_new_teacher.get(item.get("absent_teacher"), -1)
        substitute = old_to_new_teacher.get(item.get("substitute_teacher"), -1)
        if min(lesson_id, absent, substitute) < 0:
            continue
        clone = copy.deepcopy(item)
        clone.update({"lesson_id": lesson_id, "absent_teacher": absent, "substitute_teacher": substitute})
        substitutions.append(clone)
    current["substitutions"] = substitutions

    unavailable = current["teacher_unavailable"]
    managed_names = set(CONFIRMED_UNAVAILABILITY) | {
        "Елагина Ольга Александровна", "Галузин Антон Илюсович"
    }
    managed_ids = {
        teacher_id_by_name[norm(name)] for name in managed_names
        if norm(name) in teacher_id_by_name
    }
    unavailable[:] = [item for item in unavailable if item.get("teacher") not in managed_ids]
    next_unavailable = max((x.get("id", -1) for x in unavailable), default=-1) + 1
    for name, entries in CONFIRMED_UNAVAILABILITY.items():
        teacher = teacher_id_by_name.get(norm(name), -1)
        if teacher < 0:
            continue
        for entry in entries:
            unavailable.append({
                "id": next_unavailable,
                "uid": uid("teacher-unavailable", f"{teacher}:{entry}"),
                "teacher": teacher,
                "text": "Подтверждено диспетчером 31.08.2026",
                **entry,
            })
            next_unavailable += 1

    imports = current.setdefault("workload_imports", [])
    imports.append({
        "id": max((x.get("id", -1) for x in imports), default=-1) + 1,
        "file_name": source.name, "semester": 1, "imported_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "groups": len(groups), "teachers": len(teachers), "active_lessons": len(lessons),
        "mode": "replace_roster", "backup": backup.name,
    })
    current.setdefault("meta", {})["last_workload_import"] = {
        "source": str(source), "at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "vacancy_rows": vacancies,
    }

    tmp = DATA.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(current, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp.replace(DATA)
    print(json.dumps({
        "ok": True, "backup": str(backup), "groups": len(groups), "teachers": len(teachers),
        "lessons": len(lessons), "vacancy_rows": vacancies,
        "photo_schedules": sum(1 for t in teachers if t.get("availability_note")),
        "teacher_unavailable": len(unavailable), "substitutions_preserved": len(substitutions),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
