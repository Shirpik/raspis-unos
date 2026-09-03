# parse_vkleyki.py
# Парсит "Вклейки 2026-2027.xlsx" и перезаписывает data/timetable_data.json
# Берёт часы из нечётных семестров (1, 3, 5, 7) — колонка 6 каждого листа
#
# Правила total_slots:
#   УП./ВУП.  -> часы / 6   (is_block=True)
#   остальные -> часы / 2   (обычная пара, ПП, КП, экзамен)
#
# Флаги:
#   is_block = название начинается на УП. или ВУП.
#   is_pp    = название начинается на ПП.
#   is_lab   = название начинается на ЛПЗ (пробел или точка)
#   subgroup = -1 (вся группа), group_id*2 (1 подгруппа), group_id*2+1 (2 подгруппа)
#   teacher  = None если пусто или содержит «вакансия»
#
# Запуск: py parse_vkleyki.py

import openpyxl
import json
import math
import re
import os

XLSX_PATH = 'Вклейки 2026-2027.xlsx'
JSON_PATH  = os.path.join('data', 'timetable_data.json')

# ── Вспомогательные функции ──────────────────────────────────────────────────

def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def detect_subgroup(name):
    if re.search(r'1\s*подгруппа', name, re.IGNORECASE):
        return 1
    if re.search(r'2\s*подгруппа', name, re.IGNORECASE):
        return 2
    return 0  # вся группа

def clean_subgroup(name):
    name = re.sub(r'\s*\d\s*подгруппа\s*', '', name, flags=re.IGNORECASE)
    return name.strip()

def flag_lab(name):
    return bool(re.match(r'^ЛПЗ[\s.]', name, re.IGNORECASE))

def flag_block(name):
    return bool(re.match(r'^(УП|ВУП)\.', name, re.IGNORECASE))

def flag_pp(name):
    return bool(re.match(r'^ПП\.', name, re.IGNORECASE))

def is_vacancy(s):
    if not s:
        return True
    return 'вакансия' in str(s).strip().lower()

def calc_slots(hours, block):
    divisor = 6 if block else 2
    return max(1, math.ceil(hours / divisor))

# ── Subject-ID (та же логика что в import_schedule.js) ──────────────────────

_subj_by_key   = {}   # group:idx:INDEX  -> id
_subj_by_bname = {}   # group:bn:BASENAME -> id
_next_sid      = [0]

def get_subject_id(group_name, index, cleaned_name):
    # КП, УП, ВУП — нет предметного id
    if re.match(r'^(КП|УП|ВУП)[\s.]', cleaned_name, re.IGNORECASE):
        return -1
    # ПП — есть subject_id (производственная практика — полноценный предмет)

    base = re.sub(r'^ЛПЗ\s+', '', cleaned_name, flags=re.IGNORECASE).strip()
    bn_key  = f'{group_name}:bn:{base.lower()}'
    idx_key = f'{group_name}:idx:{index.strip()}' if index and index.strip() else None

    if idx_key:
        if idx_key not in _subj_by_key:
            sid = _next_sid[0]; _next_sid[0] += 1
            _subj_by_key[idx_key] = sid
            _subj_by_bname.setdefault(bn_key, sid)
        return _subj_by_key[idx_key]
    else:
        if bn_key in _subj_by_bname:
            return _subj_by_bname[bn_key]
        if bn_key not in _subj_by_key:
            sid = _next_sid[0]; _next_sid[0] += 1
            _subj_by_key[bn_key] = sid
            _subj_by_bname[bn_key] = sid
        return _subj_by_key[bn_key]

# ── Разбор листа ─────────────────────────────────────────────────────────────

def parse_sheet(ws):
    """
    Возвращает список записей из нечётного семестра (колонка 6).
    У всех листов — и 1-курсников (ауд) и 2-4 курса — нечётный семестр
    всегда оказывается в колонке 6 (проверено по реальным данным).
    """
    rows_out = []
    for r in range(11, ws.max_row + 1):
        name_raw = ws.cell(r, 3).value
        if not name_raw or not str(name_raw).strip():
            continue

        name = str(name_raw).strip()

        # Часы нечётного семестра
        hours = to_float(ws.cell(r, 6).value)
        if hours is None or hours <= 0:
            continue   # нет занятий в нечётном семестре

        rows_out.append({
            'name':        name,
            'index':       str(ws.cell(r, 2).value or '').strip(),
            'teacher_raw': ws.cell(r, 14).value,
            'total_hours': hours,
        })
    return rows_out

# ── Основной код ─────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# Сохраняем settings и unavailable из существующего JSON
with open(JSON_PATH, 'r', encoding='utf-8') as f:
    existing = json.load(f)

existing_groups     = existing['groups']
existing_settings   = existing['settings']
existing_unavailable = existing['unavailable']

# Словарь преподавателей
teacher_name_to_id = {}
teachers = []

def get_teacher_id(raw):
    if is_vacancy(raw):
        return None
    name = str(raw).strip()
    if name not in teacher_name_to_id:
        tid = len(teachers)
        teacher_name_to_id[name] = tid
        teachers.append({'id': tid, 'name': name})
    return teacher_name_to_id[name]

# Сборка уроков
lessons    = []
lesson_id  = 0
LESNAYA    = 0
KRIVOUSOVA = 1

for sheet_idx, ws in enumerate(wb.worksheets):
    if sheet_idx >= len(existing_groups):
        print(f'[!] Лист {sheet_idx} не имеет соответствующей группы — пропускаем')
        continue

    group    = existing_groups[sheet_idx]
    group_id = group['id']
    gname    = group['name']

    for entry in parse_sheet(ws):
        raw_name     = entry['name']
        index        = entry['index']
        total_hours  = entry['total_hours']
        teacher_raw  = entry['teacher_raw']

        # Определяем подгруппу
        sg_kind  = detect_subgroup(raw_name)
        cleaned  = clean_subgroup(raw_name)

        # Флаги
        lab   = flag_lab(cleaned)
        block = flag_block(cleaned)
        pp    = flag_pp(cleaned)

        # ID подгруппы: -1 = вся группа, group_id*2 = 1я, group_id*2+1 = 2я
        if sg_kind == 0:
            subgroup = -1
        else:
            subgroup = group_id * 2 + (sg_kind - 1)

        total_slots  = calc_slots(total_hours, block)
        teacher_id   = get_teacher_id(teacher_raw)
        subject_id   = get_subject_id(gname, index, cleaned)

        lessons.append({
            'id':               lesson_id,
            'group':            group_id,
            'subgroup':         subgroup,
            'teacher':          teacher_id,
            'total_hours':      int(total_hours),
            'total_slots':      total_slots,
            'name':             cleaned,
            'subject_id':       subject_id,
            'is_lab':           lab,
            'is_block':         block,
            'is_pp':            pp,
            'allowed_campuses': [LESNAYA, KRIVOUSOVA],
        })
        lesson_id += 1

output = {
    'settings':   existing_settings,
    'groups':     existing_groups,
    'teachers':   teachers,
    'unavailable': existing_unavailable,
    'lessons':    lessons,
}

os.makedirs('data', exist_ok=True)
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f'Записано: {JSON_PATH}')
print(f'  Групп:          {len(existing_groups)}')
print(f'  Преподавателей: {len(teachers)}')
print(f'  Уроков:         {len(lessons)}')

# Статистика по флагам
n_block = sum(1 for l in lessons if l['is_block'])
n_pp    = sum(1 for l in lessons if l['is_pp'])
n_lab   = sum(1 for l in lessons if l['is_lab'])
n_sub   = sum(1 for l in lessons if l['subgroup'] != -1)
n_novac = sum(1 for l in lessons if l['teacher'] is None)
print(f'  Из них:')
print(f'    is_block (УП/ВУП): {n_block}')
print(f'    is_pp    (ПП):     {n_pp}')
print(f'    is_lab   (ЛПЗ):   {n_lab}')
print(f'    подгруппные:       {n_sub}')
print(f'    без преподавателя: {n_novac}')
